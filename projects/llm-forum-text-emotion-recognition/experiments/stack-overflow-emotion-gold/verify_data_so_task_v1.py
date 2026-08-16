#!/usr/bin/env python3
"""Independently verify DATA-SO-TASK-V1 without importing its builder."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import subprocess
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PROTOCOL_ID = "DATA-SO-TASK-V1"
SOURCE_REVISION = "d6a679f39a198fdb0657a6116d35dd7b92496898"
EXPECTED_SOURCE_SHA256 = "29f667701227fc3f1ffc005c5d5364c30f24476005baac23fff8338dbd2f0179"
EXPECTED_ROWS = 4_800
EXPECTED_FRACTIONS = {"train": 0.70, "validation": 0.15, "test": 0.15}
SPLIT_ORDER = ("train", "validation", "test")
LABEL_SHEETS = (
    ("love", "Love_all"),
    ("joy", "Joy_all"),
    ("surprise", "Surprise_all"),
    ("anger", "Anger_all"),
    ("sadness", "Sadness_all"),
    ("fear", "Fear_all"),
)
LABEL_ORDER = tuple(label for label, _ in LABEL_SHEETS)
EXPECTED_LABEL_COUNTS = {
    "love": 1_220,
    "joy": 491,
    "surprise": 45,
    "anger": 882,
    "sadness": 230,
    "fear": 106,
}
EXPECTED_CARDINALITY_COUNTS = {0: 1_959, 1: 2_708, 2: 133}
FROZEN_BUCKET_QUOTAS = {
    "singleton": {"train": 3_208, "validation": 687, "test": 687},
    "duplicate-conflict-size-2": {"train": 18, "validation": 4, "test": 4},
    "duplicate-size-2": {"train": 47, "validation": 9, "test": 9},
    "duplicate-size-3": {"train": 2, "validation": 1, "test": 1},
    "duplicate-size-4": {"train": 0, "validation": 1, "test": 1},
    "duplicate-size-6": {"train": 1, "validation": 0, "test": 0},
    "duplicate-size-10": {"train": 1, "validation": 0, "test": 0},
}
SAMPLE_ID_RE = re.compile(r"^sample-[0-9a-f]{24}$")
COMPONENT_ID_RE = re.compile(r"^component-[0-9a-f]{24}$")

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]
DEFAULT_SOURCE = (
    PROJECT_ROOT
    / "data/stack-overflow-emotion-gold/official/Emotions_GoldSandard_andAnnotation.xlsx"
)
DEFAULT_PRIVATE_ROOT = (
    PROJECT_ROOT / "data/stack-overflow-emotion-gold/derived-private/task-v1"
)
DEFAULT_PUBLIC_REPORT = SCRIPT_DIR / "reports/data-so-task-v1.json"
DEFAULT_PUBLIC_MANIFEST = (
    PROJECT_ROOT / "data/stack-overflow-emotion-gold/task-v1.manifest.json"
)
DEFAULT_PUBLIC_SPLIT_INDEX = (
    PROJECT_ROOT / "data/stack-overflow-emotion-gold/task-v1.split-index.jsonl"
)
DEFAULT_VERIFICATION_REPORT = SCRIPT_DIR / "reports/data-so-task-v1-verification.json"


def digest(path: Path) -> str:
    value = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            value.update(chunk)
    return value.hexdigest()


def relative(path: Path) -> str:
    try:
        return path.resolve().relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(path.resolve())


def stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def independent_id(namespace: str, value: Any) -> str:
    payload = stable_json(
        {
            "namespace": namespace,
            "protocol_id": PROTOCOL_ID,
            "source_revision": SOURCE_REVISION,
            "value": value,
        }
    )
    return f"{namespace}-{hashlib.sha256(payload.encode('utf-8')).hexdigest()[:24]}"


def normalize_for_duplicates(text: str) -> str:
    return re.sub(
        r"\s+", " ", unicodedata.normalize("NFKC", text).casefold()
    ).strip()


def vote_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        token = value.strip()
        if not token:
            return False
        if token.lower() == "x":
            return True
    raise ValueError(f"Invalid rater value: {value!r}")


def released_gold(value: Any, label: str) -> bool:
    if value is None:
        token = ""
    elif isinstance(value, str):
        token = value.strip()
    else:
        raise ValueError(f"Invalid gold value: {value!r}")
    if not token:
        return False
    if token.lower() == label.lower():
        return True
    raise ValueError(f"Invalid gold token for {label}: {value!r}")


class DisjointSet:
    def __init__(self, count: int) -> None:
        self.parent = list(range(count))

    def root(self, value: int) -> int:
        while self.parent[value] != value:
            self.parent[value] = self.parent[self.parent[value]]
            value = self.parent[value]
        return value

    def join(self, left: int, right: int) -> None:
        left_root, right_root = self.root(left), self.root(right)
        if left_root == right_root:
            return
        low, high = sorted((left_root, right_root))
        self.parent[high] = low


def parse_source(source: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    expected_sheets = [sheet for _, sheet in LABEL_SHEETS]
    if workbook.sheetnames != expected_sheets:
        workbook.close()
        raise ValueError(f"Unexpected sheets: {workbook.sheetnames!r}")

    rows: list[dict[str, Any]] = []
    alignment_mismatches = 0
    majority_mismatches = 0
    headers_ok = True
    try:
        for sheet_index, (label, sheet_name) in enumerate(LABEL_SHEETS):
            values = list(workbook[sheet_name].iter_rows(values_only=True))
            if not values:
                raise ValueError(f"Empty sheet: {sheet_name}")
            header = values[0]
            expected_header = {
                0: "Group",
                1: "Set",
                3: "Text",
                4: "rater 1",
                5: "rater 2",
                6: "rater 3",
                7: "Gold Label",
            }
            headers_ok &= all(
                len(header) > index and header[index] == token
                for index, token in expected_header.items()
            )
            data = values[1:]
            if len(data) != EXPECTED_ROWS:
                raise ValueError(f"Unexpected row count in {sheet_name}: {len(data)}")
            for index, row in enumerate(data):
                if len(row) < 8:
                    raise ValueError(f"Short row in {sheet_name}:{index + 2}")
                group, source_set, local_number, text = row[:4]
                if not isinstance(group, str) or not group.strip():
                    raise ValueError(f"Invalid group in {sheet_name}:{index + 2}")
                if not isinstance(source_set, str) or not source_set.strip():
                    raise ValueError(f"Invalid set in {sheet_name}:{index + 2}")
                if not isinstance(local_number, int):
                    raise ValueError(f"Invalid local number in {sheet_name}:{index + 2}")
                if not isinstance(text, str) or not text.strip():
                    raise ValueError(f"Invalid text in {sheet_name}:{index + 2}")
                coordinate = (group, source_set, local_number)
                aligned = (group, source_set, local_number, text)
                if sheet_index == 0:
                    rows.append(
                        {
                            "coordinate": coordinate,
                            "aligned": aligned,
                            "text": text,
                            "labels": [],
                        }
                    )
                elif rows[index]["aligned"] != aligned:
                    alignment_mismatches += 1
                votes = [vote_value(row[column]) for column in (4, 5, 6)]
                majority = sum(votes) >= 2
                gold = released_gold(row[7], label)
                majority_mismatches += int(majority != gold)
                rows[index]["labels"].append(int(gold))
    finally:
        workbook.close()

    if not headers_ok:
        raise ValueError("One or more workbook headers do not match the protocol")
    if alignment_mismatches or majority_mismatches:
        raise ValueError(
            f"Workbook mismatch: alignment={alignment_mismatches}, majority={majority_mismatches}"
        )
    if len({row["coordinate"] for row in rows}) != len(rows):
        raise ValueError("Release coordinates are not unique")

    for row in rows:
        row["labels"] = tuple(row["labels"])
        row["neutral"] = not any(row["labels"])
        row["cardinality"] = sum(row["labels"])
        row["normalized"] = normalize_for_duplicates(row["text"])
        row["sample_id"] = independent_id(
            "sample", {"coordinate": row["coordinate"], "text": row["text"]}
        )

    exact: dict[str, list[int]] = defaultdict(list)
    normalized: dict[str, list[int]] = defaultdict(list)
    for index, row in enumerate(rows):
        exact[row["text"]].append(index)
        normalized[row["normalized"]].append(index)
    union = DisjointSet(len(rows))
    for groups in (exact, normalized):
        for indices in groups.values():
            for index in indices[1:]:
                union.join(indices[0], index)
    grouped: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for index, row in enumerate(rows):
        grouped[union.root(index)].append(row)
    components: list[dict[str, Any]] = []
    for members in grouped.values():
        component_id = independent_id(
            "component", sorted(row["sample_id"] for row in members)
        )
        conflicting = len({row["labels"] for row in members}) > 1
        for row in members:
            row["component_id"] = component_id
        components.append(
            {
                "component_id": component_id,
                "members": members,
                "size": len(members),
                "conflicting": conflicting,
            }
        )

    def group_summary(groups: dict[str, list[int]]) -> dict[str, int]:
        duplicates = [indices for indices in groups.values() if len(indices) > 1]
        conflicts = [
            indices
            for indices in duplicates
            if len({rows[index]["labels"] for index in indices}) > 1
        ]
        return {
            "unique_keys": len(groups),
            "duplicate_components": len(duplicates),
            "duplicate_rows": sum(len(indices) for indices in duplicates),
            "extra_duplicate_rows": sum(len(indices) - 1 for indices in duplicates),
            "maximum_component_size": max((len(indices) for indices in duplicates), default=1),
            "conflicting_components": len(conflicts),
            "conflicting_rows": sum(len(indices) for indices in conflicts),
        }

    audit = {
        "exact": group_summary(exact),
        "normalized": group_summary(normalized),
        "connected_components": len(components),
        "conflicting_connected_components": sum(
            component["conflicting"] for component in components
        ),
        "conflicting_connected_component_rows": sum(
            component["size"] for component in components if component["conflicting"]
        ),
        "components": components,
    }
    return rows, audit


def component_bucket(component: dict[str, Any]) -> str:
    if component["size"] == 1:
        return "singleton"
    if component["conflicting"]:
        return f"duplicate-conflict-size-{component['size']}"
    return f"duplicate-size-{component['size']}"


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    values: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, 1):
            if not line.strip():
                raise ValueError(f"Blank line in {path}:{line_number}")
            value = json.loads(line)
            if not isinstance(value, dict):
                raise ValueError(f"Non-object JSONL row in {path}:{line_number}")
            values.append(value)
    return values


def recursive_keys(value: Any) -> set[str]:
    found: set[str] = set()
    if isinstance(value, dict):
        found.update(str(key) for key in value)
        for nested in value.values():
            found.update(recursive_keys(nested))
    elif isinstance(value, list):
        for nested in value:
            found.update(recursive_keys(nested))
    return found


def recursive_strings(value: Any) -> set[str]:
    found: set[str] = set()
    if isinstance(value, str):
        found.add(value)
    elif isinstance(value, dict):
        for key, nested in value.items():
            found.add(str(key))
            found.update(recursive_strings(nested))
    elif isinstance(value, list):
        for nested in value:
            found.update(recursive_strings(nested))
    return found


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def file_record(path: Path) -> dict[str, Any]:
    return {"path": relative(path), "bytes": path.stat().st_size, "sha256": digest(path)}


class Checks:
    def __init__(self) -> None:
        self.results: list[dict[str, Any]] = []

    def add(self, name: str, passed: bool, detail: Any = None) -> None:
        self.results.append({"name": name, "passed": bool(passed), "detail": detail})

    @property
    def failures(self) -> list[dict[str, Any]]:
        return [result for result in self.results if not result["passed"]]


def verify(
    source: Path,
    private_root: Path,
    public_report_path: Path,
    public_manifest_path: Path,
    public_split_index_path: Path,
    verification_report_path: Path,
) -> dict[str, Any]:
    checks = Checks()
    source_digest = digest(source)
    checks.add("source_sha256", source_digest == EXPECTED_SOURCE_SHA256, source_digest)
    rows, audit = parse_source(source)
    components = audit.pop("components")

    label_counts = {
        label: sum(row["labels"][index] for row in rows)
        for index, label in enumerate(LABEL_ORDER)
    }
    cardinality_counts = Counter(row["cardinality"] for row in rows)
    checks.add("row_count", len(rows) == EXPECTED_ROWS, len(rows))
    checks.add("sample_ids_unique", len({row["sample_id"] for row in rows}) == len(rows))
    checks.add("sample_id_format", all(SAMPLE_ID_RE.fullmatch(row["sample_id"]) for row in rows))
    checks.add("label_counts", label_counts == EXPECTED_LABEL_COUNTS, label_counts)
    checks.add(
        "cardinality_counts",
        dict(sorted(cardinality_counts.items())) == EXPECTED_CARDINALITY_COUNTS,
        dict(cardinality_counts),
    )
    expected_duplicate_audit = {
        "exact": {
            "unique_keys": 4_687,
            "duplicate_components": 99,
            "duplicate_rows": 212,
            "extra_duplicate_rows": 113,
            "maximum_component_size": 7,
            "conflicting_components": 26,
            "conflicting_rows": 52,
        },
        "normalized": {
            "unique_keys": 4_681,
            "duplicate_components": 99,
            "duplicate_rows": 218,
            "extra_duplicate_rows": 119,
            "maximum_component_size": 10,
            "conflicting_components": 26,
            "conflicting_rows": 52,
        },
        "connected_components": 4_681,
        "conflicting_connected_components": 26,
        "conflicting_connected_component_rows": 52,
    }
    checks.add("duplicate_audit", audit == expected_duplicate_audit, audit)
    checks.add(
        "component_id_format",
        all(COMPONENT_ID_RE.fullmatch(component["component_id"]) for component in components),
    )

    train = load_jsonl(private_root / "train.jsonl")
    validation = load_jsonl(private_root / "validation.jsonl")
    test_inputs = load_jsonl(private_root / "test.inputs.jsonl")
    test_labels = load_jsonl(private_root / "test.labels.sealed.jsonl")
    conflict_rows = load_jsonl(private_root / "duplicate-conflicts.jsonl")
    private_manifest = json.loads(
        (private_root / "private-manifest.json").read_text(encoding="utf-8")
    )
    public_report = json.loads(public_report_path.read_text(encoding="utf-8"))
    public_manifest = json.loads(public_manifest_path.read_text(encoding="utf-8"))
    public_split_index = load_jsonl(public_split_index_path)

    labeled_schema = {
        "schema_version",
        "protocol_id",
        "sample_id",
        "component_id",
        "text",
        "labels",
        "neutral",
        "label_cardinality",
    }
    test_input_schema = {
        "schema_version",
        "protocol_id",
        "sample_id",
        "component_id",
        "text",
    }
    sealed_schema = {"schema_version", "protocol_id", "sample_id", "labels"}
    split_index_schema = {
        "schema_version",
        "protocol_id",
        "sample_id",
        "component_id",
        "split",
    }
    checks.add(
        "train_validation_schema",
        all(set(value) == labeled_schema for value in train + validation),
    )
    checks.add("test_input_minimal_schema", all(set(value) == test_input_schema for value in test_inputs))
    checks.add("sealed_label_minimal_schema", all(set(value) == sealed_schema for value in test_labels))
    checks.add(
        "public_split_index_minimal_schema",
        all(set(value) == split_index_schema for value in public_split_index),
    )
    checks.add(
        "protocol_ids",
        all(
            value.get("protocol_id") == PROTOCOL_ID
            for value in train
            + validation
            + test_inputs
            + test_labels
            + conflict_rows
            + public_split_index
        ),
    )

    sealed_by_id = {value["sample_id"]: value for value in test_labels}
    checks.add("sealed_ids_unique", len(sealed_by_id) == len(test_labels))
    checks.add(
        "test_input_and_label_ids_match",
        {value["sample_id"] for value in test_inputs} == set(sealed_by_id),
    )
    expected_by_id = {row["sample_id"]: row for row in rows}
    actual_by_id: dict[str, tuple[str, dict[str, Any]]] = {}
    duplicate_actual_ids = 0
    for split, values in (
        ("train", train),
        ("validation", validation),
        ("test", test_inputs),
    ):
        for value in values:
            sample_id = value["sample_id"]
            duplicate_actual_ids += int(sample_id in actual_by_id)
            actual_by_id[sample_id] = (split, value)
    checks.add("private_sample_ids_unique", duplicate_actual_ids == 0, duplicate_actual_ids)
    checks.add(
        "private_record_set_exact",
        set(actual_by_id) == set(expected_by_id),
        {"expected": len(expected_by_id), "actual": len(actual_by_id)},
    )

    field_mismatches = 0
    split_by_component: dict[str, set[str]] = defaultdict(set)
    split_by_exact: dict[str, set[str]] = defaultdict(set)
    split_by_normalized: dict[str, set[str]] = defaultdict(set)
    split_rows: Counter[str] = Counter()
    split_labels: dict[str, Counter[str]] = defaultdict(Counter)
    split_slices: dict[str, Counter[str]] = defaultdict(Counter)
    assignment_by_sample: dict[str, str] = {}
    for sample_id, (split, value) in actual_by_id.items():
        expected = expected_by_id.get(sample_id)
        if expected is None:
            field_mismatches += 1
            continue
        labels = (
            tuple(sealed_by_id[sample_id]["labels"])
            if split == "test"
            else tuple(value.get("labels", ()))
        )
        observed = {
            "component_id": value.get("component_id"),
            "text": value.get("text"),
            "labels": labels,
        }
        reference = {
            "component_id": expected["component_id"],
            "text": expected["text"],
            "labels": expected["labels"],
        }
        field_mismatches += int(observed != reference)
        if split != "test":
            field_mismatches += int(value.get("neutral") is not expected["neutral"])
            field_mismatches += int(
                value.get("label_cardinality") != expected["cardinality"]
            )
        split_by_component[expected["component_id"]].add(split)
        split_by_exact[expected["text"]].add(split)
        split_by_normalized[expected["normalized"]].add(split)
        assignment_by_sample[sample_id] = split
        split_rows[split] += 1
        for label, present in zip(LABEL_ORDER, labels):
            split_labels[split][label] += present
        split_slices[split]["neutral"] += int(not any(labels))
        split_slices[split][f"cardinality_{sum(labels)}"] += 1
    checks.add("private_fields_exact", field_mismatches == 0, field_mismatches)
    checks.add(
        "components_do_not_cross_splits",
        all(len(splits) == 1 for splits in split_by_component.values()),
    )
    checks.add(
        "exact_text_does_not_cross_splits",
        all(len(splits) == 1 for splits in split_by_exact.values()),
    )
    checks.add(
        "normalized_text_does_not_cross_splits",
        all(len(splits) == 1 for splits in split_by_normalized.values()),
    )

    index_by_id = {value["sample_id"]: value for value in public_split_index}
    checks.add("public_split_ids_unique", len(index_by_id) == len(public_split_index))
    checks.add("public_split_record_set_exact", set(index_by_id) == set(expected_by_id))
    checks.add(
        "public_split_assignments_exact",
        all(
            value.get("split") == assignment_by_sample.get(sample_id)
            and value.get("component_id") == expected_by_id[sample_id]["component_id"]
            for sample_id, value in index_by_id.items()
        ),
    )

    row_errors = {
        split: abs(split_rows[split] / len(rows) - fraction)
        for split, fraction in EXPECTED_FRACTIONS.items()
    }
    checks.add("split_row_ratio_tolerance", max(row_errors.values()) <= 0.005, row_errors)
    checks.add(
        "all_labels_present_in_every_split",
        all(split_labels[split][label] > 0 for split in SPLIT_ORDER for label in LABEL_ORDER),
        {split: dict(split_labels[split]) for split in SPLIT_ORDER},
    )
    label_errors: dict[str, float] = {}
    slice_errors: dict[str, float] = {}
    for split, fraction in EXPECTED_FRACTIONS.items():
        for label in LABEL_ORDER:
            label_errors[f"{split}:{label}"] = abs(
                split_labels[split][label] / EXPECTED_LABEL_COUNTS[label] - fraction
            )
        for name, total in (
            ("neutral", EXPECTED_CARDINALITY_COUNTS[0]),
            ("cardinality_1", EXPECTED_CARDINALITY_COUNTS[1]),
            ("cardinality_2", EXPECTED_CARDINALITY_COUNTS[2]),
        ):
            slice_errors[f"{split}:{name}"] = abs(
                split_slices[split][name] / total - fraction
            )
    checks.add(
        "label_allocation_tolerance",
        max(label_errors.values()) <= 0.05,
        {"max_absolute_error": round(max(label_errors.values()), 6)},
    )
    checks.add(
        "neutral_cardinality_tolerance",
        max(slice_errors.values()) <= 0.03,
        {"max_absolute_error": round(max(slice_errors.values()), 6)},
    )

    bucket_counts: dict[str, Counter[str]] = defaultdict(Counter)
    for component in components:
        splits = split_by_component[component["component_id"]]
        if len(splits) == 1:
            bucket_counts[component_bucket(component)][next(iter(splits))] += 1
    observed_quotas = {
        bucket: {split: bucket_counts[bucket][split] for split in SPLIT_ORDER}
        for bucket in FROZEN_BUCKET_QUOTAS
    }
    checks.add("frozen_component_bucket_quotas", observed_quotas == FROZEN_BUCKET_QUOTAS, observed_quotas)

    expected_conflicts = []
    for component in components:
        if not component["conflicting"]:
            continue
        expected_conflicts.append(
            {
                "schema_version": "so-emotion-duplicate-conflict-v1",
                "protocol_id": PROTOCOL_ID,
                "component_id": component["component_id"],
                "members": [
                    {"sample_id": row["sample_id"], "labels": list(row["labels"])}
                    for row in sorted(component["members"], key=lambda value: value["sample_id"])
                ],
            }
        )
    expected_conflicts.sort(key=lambda value: value["component_id"])
    checks.add("private_conflict_diagnostics_exact", conflict_rows == expected_conflicts)

    checks.add(
        "public_protocol_ids",
        public_report.get("protocol_id") == PROTOCOL_ID
        and public_manifest.get("protocol_id") == PROTOCOL_ID,
    )
    checks.add(
        "public_source_metadata",
        public_report.get("source", {}).get("revision") == SOURCE_REVISION
        and public_manifest.get("source", {}).get("revision") == SOURCE_REVISION
        and public_manifest.get("source", {}).get("sha256") == source_digest,
    )
    checks.add(
        "public_aggregate_counts",
        public_report.get("construction", {}).get("label_counts") == label_counts
        and public_report.get("duplicates") == audit
        and public_manifest.get("counts", {}).get("rows") == len(rows)
        and public_manifest.get("counts", {}).get("components") == len(components)
        and public_manifest.get("counts", {}).get("split_rows")
        == {split: split_rows[split] for split in SPLIT_ORDER},
    )
    checks.add(
        "public_split_index_hash",
        public_manifest.get("public_split_index", {}).get("sha256")
        == digest(public_split_index_path),
    )
    checks.add(
        "public_report_hash",
        public_manifest.get("public_report", {}).get("sha256")
        == digest(public_report_path),
    )

    artifact_paths = {
        "train": private_root / "train.jsonl",
        "validation": private_root / "validation.jsonl",
        "test_inputs": private_root / "test.inputs.jsonl",
        "test_labels": private_root / "test.labels.sealed.jsonl",
        "duplicate_conflicts": private_root / "duplicate-conflicts.jsonl",
        "private_manifest": private_root / "private-manifest.json",
    }
    manifest_artifacts = public_manifest.get("private_artifacts", {})
    for name, path in artifact_paths.items():
        checks.add(
            f"private_artifact_hash:{name}",
            manifest_artifacts.get(name, {}).get("sha256") == digest(path),
            digest(path),
        )
    private_file_modes = {
        name: oct(path.stat().st_mode & 0o777) for name, path in artifact_paths.items()
    }
    checks.add(
        "private_files_owner_only",
        all((path.stat().st_mode & 0o077) == 0 for path in artifact_paths.values()),
        private_file_modes,
    )
    checks.add(
        "private_directory_owner_only",
        (private_root.stat().st_mode & 0o077) == 0,
        oct(private_root.stat().st_mode & 0o777),
    )
    checks.add(
        "private_parent_directory_owner_only",
        (private_root.parent.stat().st_mode & 0o077) == 0,
        oct(private_root.parent.stat().st_mode & 0o777),
    )
    checks.add(
        "official_source_owner_only",
        (source.stat().st_mode & 0o077) == 0,
        oct(source.stat().st_mode & 0o777),
    )
    checks.add(
        "official_directory_owner_only",
        (source.parent.stat().st_mode & 0o077) == 0,
        oct(source.parent.stat().st_mode & 0o777),
    )
    checks.add(
        "private_manifest_metadata",
        private_manifest.get("protocol_id") == PROTOCOL_ID
        and private_manifest.get("source_revision") == SOURCE_REVISION
        and private_manifest.get("source_sha256") == source_digest
        and private_manifest.get("test_gate", {}).get("status")
        == "sealed_not_authorized_for_model_access",
    )

    raw_texts = {row["text"] for row in rows}
    public_row_keys = {"text", "labels", "rater_votes", "source_coordinate"}
    public_aggregates = [public_report, public_manifest]
    checks.add(
        "public_reports_have_no_row_level_keys",
        all(not (recursive_keys(value) & public_row_keys) for value in public_aggregates),
    )
    public_strings = set().union(*(recursive_strings(value) for value in public_aggregates))
    checks.add(
        "public_reports_have_no_exact_raw_text_values",
        not (raw_texts & public_strings),
    )
    checks.add(
        "public_split_index_has_no_gold_or_text",
        all(
            not (set(value) & {"text", "labels", "neutral", "label_cardinality", "gold", "rater_votes"})
            for value in public_split_index
        ),
    )

    private_probe = DEFAULT_PRIVATE_ROOT / "__ignore_probe__.txt"
    source_probe = DEFAULT_SOURCE
    public_probe = DEFAULT_PUBLIC_SPLIT_INDEX
    private_ignored = subprocess.run(
        ["git", "check-ignore", "-q", str(private_probe)],
        cwd=PROJECT_ROOT,
        check=False,
        capture_output=True,
    ).returncode == 0
    source_ignored = subprocess.run(
        ["git", "check-ignore", "-q", str(source_probe)],
        cwd=PROJECT_ROOT,
        check=False,
        capture_output=True,
    ).returncode == 0
    public_ignored = subprocess.run(
        ["git", "check-ignore", "-q", str(public_probe)],
        cwd=PROJECT_ROOT,
        check=False,
        capture_output=True,
    ).returncode == 0
    checks.add("private_root_gitignored", private_ignored)
    checks.add("official_source_gitignored", source_ignored)
    checks.add("public_split_index_not_gitignored", not public_ignored)

    verification = {
        "schema_version": "so-emotion-c0-verification-report-v1",
        "protocol_id": PROTOCOL_ID,
        "status": "verified" if not checks.failures else "failed",
        "checks_total": len(checks.results),
        "checks_passed": len(checks.results) - len(checks.failures),
        "checks_failed": len(checks.failures),
        "failures": checks.failures,
        "aggregate": {
            "rows": len(rows),
            "components": len(components),
            "label_counts": label_counts,
            "cardinality_counts": {
                str(key): value for key, value in sorted(cardinality_counts.items())
            },
            "split_rows": {split: split_rows[split] for split in SPLIT_ORDER},
            "frozen_bucket_quotas": observed_quotas,
            "maximum_label_allocation_error": round(max(label_errors.values()), 6),
            "maximum_neutral_or_cardinality_error": round(max(slice_errors.values()), 6),
            "duplicate_audit": audit,
        },
        "test_gate": "sealed_not_authorized_for_model_access",
        "claim_boundary": (
            "This verifies data construction, leakage controls, privacy boundaries and "
            "test sealing only; no model result was produced or inspected."
        ),
    }
    write_json(verification_report_path, verification)
    if verification["status"] == "verified":
        public_report["status"] = "verified"
        public_report["independent_verification"] = {
            "status": "verified",
            "checks_passed": verification["checks_passed"],
            "checks_total": verification["checks_total"],
            "report": relative(verification_report_path),
        }
        write_json(public_report_path, public_report)
        public_manifest["status"] = "verified"
        public_manifest["public_report"] = file_record(public_report_path)
        public_manifest["verification_report"] = file_record(verification_report_path)
        write_json(public_manifest_path, public_manifest)
    return verification


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--private-root", type=Path, default=DEFAULT_PRIVATE_ROOT)
    parser.add_argument("--public-report", type=Path, default=DEFAULT_PUBLIC_REPORT)
    parser.add_argument("--public-manifest", type=Path, default=DEFAULT_PUBLIC_MANIFEST)
    parser.add_argument(
        "--public-split-index", type=Path, default=DEFAULT_PUBLIC_SPLIT_INDEX
    )
    parser.add_argument(
        "--verification-report", type=Path, default=DEFAULT_VERIFICATION_REPORT
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        result = verify(
            args.source.resolve(),
            args.private_root.resolve(),
            args.public_report.resolve(),
            args.public_manifest.resolve(),
            args.public_split_index.resolve(),
            args.verification_report.resolve(),
        )
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1
    print(
        stable_json(
            {
                "protocol_id": PROTOCOL_ID,
                "status": result["status"],
                "checks_passed": result["checks_passed"],
                "checks_total": result["checks_total"],
            }
        )
    )
    return 0 if result["status"] == "verified" else 1


if __name__ == "__main__":
    raise SystemExit(main())

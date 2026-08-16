#!/usr/bin/env python3
"""Build the private Stack Overflow C0 task and public aggregate artifacts."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import random
import re
import shutil
import statistics
import subprocess
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PROTOCOL_ID = "DATA-SO-TASK-V1"
SCHEMA_VERSION = "so-emotion-c0-v1"
FROZEN_DATE = "2026-08-13"
SEED = 20260813
SOURCE_REVISION = "d6a679f39a198fdb0657a6116d35dd7b92496898"
EXPECTED_SOURCE_SHA256 = "29f667701227fc3f1ffc005c5d5364c30f24476005baac23fff8338dbd2f0179"
EXPECTED_ROWS = 4_800
SPLIT_FRACTIONS = {"train": 0.70, "validation": 0.15, "test": 0.15}
SPLIT_ORDER = ("train", "validation", "test")
LABEL_SHEETS = (
    ("love", "Love_all"),
    ("joy", "Joy_all"),
    ("surprise", "Surprise_all"),
    ("anger", "Anger_all"),
    ("sadness", "Sadness_all"),
    ("fear", "Fear_all"),
)
LABEL_ORDER = tuple(label for label, _ in LABEL_SHEETS)
EXPECTED_LABEL_COUNTS = {
    "love": 1_220,
    "joy": 491,
    "surprise": 45,
    "anger": 882,
    "sadness": 230,
    "fear": 106,
}
EXPECTED_CARDINALITY_COUNTS = {0: 1_959, 1: 2_708, 2: 133}
BALANCE_SLICES = (
    "neutral",
    "cardinality_1",
    "cardinality_2",
    "component_count",
    "duplicate_rows",
    "duplicate_components",
    "conflicting_duplicate_rows",
    "conflicting_duplicate_components",
)
STRATA_ORDER = LABEL_ORDER + BALANCE_SLICES
FROZEN_BUCKET_QUOTAS = {
    "singleton": {"train": 3_208, "validation": 687, "test": 687},
    "duplicate-conflict-size-2": {"train": 18, "validation": 4, "test": 4},
    "duplicate-size-2": {"train": 47, "validation": 9, "test": 9},
    "duplicate-size-3": {"train": 2, "validation": 1, "test": 1},
    "duplicate-size-4": {"train": 0, "validation": 1, "test": 1},
    "duplicate-size-6": {"train": 1, "validation": 0, "test": 0},
    "duplicate-size-10": {"train": 1, "validation": 0, "test": 0},
}
BUCKET_ORDER = (
    "duplicate-size-10",
    "duplicate-size-6",
    "duplicate-size-4",
    "duplicate-size-3",
    "duplicate-conflict-size-2",
    "duplicate-size-2",
    "singleton",
)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]
DEFAULT_SOURCE = (
    PROJECT_ROOT
    / "data/stack-overflow-emotion-gold/official/Emotions_GoldSandard_andAnnotation.xlsx"
)
DEFAULT_PRIVATE_ROOT = (
    PROJECT_ROOT / "data/stack-overflow-emotion-gold/derived-private/task-v1"
)
DEFAULT_PUBLIC_REPORT = SCRIPT_DIR / "reports/data-so-task-v1.json"
DEFAULT_PUBLIC_MANIFEST = (
    PROJECT_ROOT / "data/stack-overflow-emotion-gold/task-v1.manifest.json"
)
DEFAULT_PUBLIC_SPLIT_INDEX = (
    PROJECT_ROOT / "data/stack-overflow-emotion-gold/task-v1.split-index.jsonl"
)


def sha256_file(path: Path) -> str:
    value = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            value.update(chunk)
    return value.hexdigest()


def relative(path: Path) -> str:
    try:
        return path.resolve().relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(path.resolve())


def stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def opaque_id(namespace: str, value: Any) -> str:
    payload = stable_json(
        {
            "namespace": namespace,
            "protocol_id": PROTOCOL_ID,
            "source_revision": SOURCE_REVISION,
            "value": value,
        }
    )
    return f"{namespace}-{hashlib.sha256(payload.encode('utf-8')).hexdigest()[:24]}"


def normalized_text(text: str) -> str:
    value = unicodedata.normalize("NFKC", text).casefold()
    return re.sub(r"\s+", " ", value).strip()


def marked(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        token = value.strip()
        if not token:
            return False
        if token.casefold() == "x":
            return True
    raise ValueError(f"Unexpected rater cell value: {value!r}")


def gold_present(value: Any, label: str) -> bool:
    if value is None:
        token = ""
    elif isinstance(value, str):
        token = value.strip()
    else:
        raise ValueError(f"Unexpected Gold Label cell value for {label}: {value!r}")
    if not token:
        return False
    if token.casefold() == label.casefold():
        return True
    raise ValueError(f"Unexpected Gold Label value for {label}: {value!r}")


def validate_header(sheet_name: str, row: tuple[Any, ...]) -> None:
    expected = {
        0: "Group",
        1: "Set",
        3: "Text",
        4: "rater 1",
        5: "rater 2",
        6: "rater 3",
        7: "Gold Label",
    }
    for index, value in expected.items():
        if index >= len(row) or value != expected[index]:
            raise ValueError(
                f"Unexpected header in {sheet_name} column {index + 1}: {value!r}"
            )


def load_records(source: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    expected_sheets = [sheet for _, sheet in LABEL_SHEETS]
    if workbook.sheetnames != expected_sheets:
        workbook.close()
        raise ValueError(
            f"Unexpected sheets: expected {expected_sheets!r}, got {workbook.sheetnames!r}"
        )

    records: list[dict[str, Any]] = []
    try:
        for label_index, (label, sheet_name) in enumerate(LABEL_SHEETS):
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            try:
                header = next(iterator)
            except StopIteration as error:
                raise ValueError(f"Empty sheet: {sheet_name}") from error
            validate_header(sheet_name, header)
            rows = list(iterator)
            if len(rows) != EXPECTED_ROWS:
                raise ValueError(
                    f"Unexpected row count in {sheet_name}: {len(rows)} != {EXPECTED_ROWS}"
                )

            for row_index, row in enumerate(rows):
                if len(row) < 8:
                    raise ValueError(f"Short row in {sheet_name}:{row_index + 2}")
                group, source_set, local_number, text = row[:4]
                if not isinstance(group, str) or not group.strip():
                    raise ValueError(f"Invalid Group in {sheet_name}:{row_index + 2}")
                if not isinstance(source_set, str) or not source_set.strip():
                    raise ValueError(f"Invalid Set in {sheet_name}:{row_index + 2}")
                if not isinstance(local_number, int):
                    raise ValueError(
                        f"Invalid local number in {sheet_name}:{row_index + 2}: {local_number!r}"
                    )
                if not isinstance(text, str) or not text.strip():
                    raise ValueError(f"Invalid Text in {sheet_name}:{row_index + 2}")

                source_coordinate = (group, source_set, local_number)
                alignment = (group, source_set, local_number, text)
                if label_index == 0:
                    records.append(
                        {
                            "row_index": row_index,
                            "source_coordinate": source_coordinate,
                            "alignment": alignment,
                            "text": text,
                            "rater_votes": [],
                            "labels": [],
                        }
                    )
                elif records[row_index]["alignment"] != alignment:
                    raise ValueError(
                        f"Sheet alignment mismatch in {sheet_name}:{row_index + 2}"
                    )

                votes = tuple(marked(row[index]) for index in (4, 5, 6))
                majority = sum(votes) >= 2
                gold = gold_present(row[7], label)
                if majority != gold:
                    raise ValueError(
                        f"Majority/gold mismatch for {label} in row {row_index + 2}"
                    )
                records[row_index]["rater_votes"].append(votes)
                records[row_index]["labels"].append(int(gold))
    finally:
        workbook.close()

    coordinates = [record["source_coordinate"] for record in records]
    if len(set(coordinates)) != len(coordinates):
        raise ValueError("Release coordinates are not unique")

    for record in records:
        record["labels"] = tuple(record["labels"])
        record["rater_votes"] = tuple(record["rater_votes"])
        record["neutral"] = not any(record["labels"])
        record["label_cardinality"] = sum(record["labels"])
        record["normalized_text"] = normalized_text(record["text"])
        if not record["normalized_text"]:
            raise ValueError("Text is empty after duplicate-only normalization")
        record["sample_id"] = opaque_id(
            "sample",
            {
                "coordinate": record["source_coordinate"],
                "text": record["text"],
            },
        )
    if len({record["sample_id"] for record in records}) != len(records):
        raise ValueError("Opaque sample IDs are not unique")
    return records


class UnionFind:
    def __init__(self, values: Iterable[int]) -> None:
        self.parent = {value: value for value in values}

    def find(self, value: int) -> int:
        parent = self.parent[value]
        if parent != value:
            self.parent[value] = self.find(parent)
        return self.parent[value]

    def union(self, left: int, right: int) -> None:
        left_root = self.find(left)
        right_root = self.find(right)
        if left_root == right_root:
            return
        low, high = sorted((left_root, right_root))
        self.parent[high] = low


def duplicate_groups(records: list[dict[str, Any]], key: str) -> dict[str, list[int]]:
    groups: dict[str, list[int]] = defaultdict(list)
    for index, record in enumerate(records):
        groups[record[key]].append(index)
    return groups


def duplicate_summary(
    groups: dict[str, list[int]], records: list[dict[str, Any]]
) -> dict[str, int]:
    duplicates = [indices for indices in groups.values() if len(indices) > 1]
    conflicts = [
        indices
        for indices in duplicates
        if len({records[index]["labels"] for index in indices}) > 1
    ]
    return {
        "unique_keys": len(groups),
        "duplicate_components": len(duplicates),
        "duplicate_rows": sum(len(indices) for indices in duplicates),
        "extra_duplicate_rows": sum(len(indices) - 1 for indices in duplicates),
        "maximum_component_size": max((len(indices) for indices in duplicates), default=1),
        "conflicting_components": len(conflicts),
        "conflicting_rows": sum(len(indices) for indices in conflicts),
    }


def build_components(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    exact_groups = duplicate_groups(records, "text")
    normalized_groups = duplicate_groups(records, "normalized_text")
    union_find = UnionFind(range(len(records)))
    for groups in (exact_groups, normalized_groups):
        for indices in groups.values():
            first = indices[0]
            for index in indices[1:]:
                union_find.union(first, index)

    by_root: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for index, record in enumerate(records):
        by_root[union_find.find(index)].append(record)

    components: list[dict[str, Any]] = []
    for members in by_root.values():
        sample_ids = sorted(record["sample_id"] for record in members)
        component_id = opaque_id("component", sample_ids)
        strata: Counter[str] = Counter()
        for record in members:
            for label, present in zip(LABEL_ORDER, record["labels"]):
                strata[label] += present
            strata["neutral"] += int(record["neutral"])
            strata[f"cardinality_{record['label_cardinality']}"] += 1
            record["component_id"] = component_id
        conflicting_labels = len({record["labels"] for record in members}) > 1
        strata["component_count"] = 1
        strata["duplicate_rows"] = len(members) if len(members) > 1 else 0
        strata["duplicate_components"] = int(len(members) > 1)
        strata["conflicting_duplicate_rows"] = (
            len(members) if conflicting_labels else 0
        )
        strata["conflicting_duplicate_components"] = int(conflicting_labels)
        components.append(
            {
                "component_id": component_id,
                "records": members,
                "row_count": len(members),
                "strata": strata,
                "conflicting_labels": conflicting_labels,
            }
        )
    components.sort(key=lambda value: value["component_id"])
    if len({value["component_id"] for value in components}) != len(components):
        raise ValueError("Opaque component IDs are not unique")
    audit = {
        "exact": duplicate_summary(exact_groups, records),
        "normalized": duplicate_summary(normalized_groups, records),
        "connected_components": len(components),
        "conflicting_connected_components": sum(
            component["conflicting_labels"] for component in components
        ),
        "conflicting_connected_component_rows": sum(
            component["row_count"]
            for component in components
            if component["conflicting_labels"]
        ),
    }
    return components, audit


def total_strata(components: list[dict[str, Any]]) -> Counter[str]:
    value: Counter[str] = Counter()
    for component in components:
        value.update(component["strata"])
    return value


def state_cost(
    split: str,
    rows: int,
    strata: Counter[str],
    total_rows: int,
    totals: Counter[str],
) -> float:
    fraction = SPLIT_FRACTIONS[split]
    row_target = total_rows * fraction
    cost = 4096.0 * ((rows - row_target) / max(row_target, 1.0)) ** 2
    for name in STRATA_ORDER:
        target = totals[name] * fraction
        if name == "surprise":
            weight = 64.0
        elif name in LABEL_ORDER:
            weight = 12.0
        elif name in {
            "duplicate_components",
            "conflicting_duplicate_rows",
            "conflicting_duplicate_components",
        }:
            weight = 64.0
        else:
            weight = 24.0
        cost += weight * ((strata[name] - target) / max(target, 2.0)) ** 2
    return cost


def allocation_bucket(component: dict[str, Any]) -> str:
    if component["row_count"] == 1:
        return "singleton"
    if component["conflicting_labels"]:
        return f"duplicate-conflict-size-{component['row_count']}"
    return f"duplicate-size-{component['row_count']}"


def allocate_splits(components: list[dict[str, Any]]) -> dict[str, str]:
    total_rows = sum(component["row_count"] for component in components)
    totals = total_strata(components)
    assignments: dict[str, str] = {}
    states: dict[str, dict[str, Any]] = {
        split: {"rows": 0, "strata": Counter()} for split in SPLIT_ORDER
    }

    by_bucket: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for component in components:
        by_bucket[allocation_bucket(component)].append(component)
    observed_buckets = {name: len(values) for name, values in by_bucket.items()}
    expected_buckets = {
        name: sum(quotas.values()) for name, quotas in FROZEN_BUCKET_QUOTAS.items()
    }
    if observed_buckets != expected_buckets:
        raise RuntimeError(
            f"Frozen component bucket counts changed: {observed_buckets} != {expected_buckets}"
        )

    for bucket in BUCKET_ORDER:
        remaining_quota = FROZEN_BUCKET_QUOTAS[bucket].copy()
        values = sorted(
            by_bucket[bucket],
            key=lambda component: (
                -sum(
                    component["strata"][label] / max(totals[label], 1)
                    for label in LABEL_ORDER
                ),
                hashlib.sha256(
                    f"{SEED}:{bucket}:{component['component_id']}".encode("ascii")
                ).hexdigest(),
            ),
        )
        for component in values:
            candidates: list[tuple[float, str, str]] = []
            for split in SPLIT_ORDER:
                if remaining_quota[split] <= 0:
                    continue
                after_rows = states[split]["rows"] + component["row_count"]
                after_strata = states[split]["strata"] + component["strata"]
                score = state_cost(split, after_rows, after_strata, total_rows, totals)
                tie = hashlib.sha256(
                    f"{SEED}:{bucket}:{split}:{component['component_id']}".encode("ascii")
                ).hexdigest()
                candidates.append((score, tie, split))
            if not candidates:
                raise RuntimeError(f"No remaining quota for {bucket}")
            _, _, chosen_split = min(candidates)
            assignments[component["component_id"]] = chosen_split
            remaining_quota[chosen_split] -= 1
            states[chosen_split]["rows"] += component["row_count"]
            states[chosen_split]["strata"].update(component["strata"])
        if any(remaining_quota.values()):
            raise RuntimeError(f"Unfilled frozen quota for {bucket}: {remaining_quota}")

    expected_rows = {
        split: round(total_rows * fraction)
        for split, fraction in SPLIT_FRACTIONS.items()
    }
    actual_rows = {split: states[split]["rows"] for split in SPLIT_ORDER}
    if actual_rows != expected_rows:
        raise RuntimeError(f"Frozen bucket quotas do not produce exact rows: {actual_rows}")

    def swapped(
        current: Counter[str], remove: Counter[str], add: Counter[str]
    ) -> Counter[str]:
        value = current.copy()
        value.subtract(remove)
        value.update(add)
        return value

    rng = random.Random(SEED)
    refinable_buckets = [
        bucket
        for bucket in BUCKET_ORDER
        if sum(value > 0 for value in FROZEN_BUCKET_QUOTAS[bucket].values()) > 1
    ]
    iterations = min(450_000, max(60_000, len(components) * 100))
    for _ in range(iterations):
        bucket = rng.choice(refinable_buckets)
        left, right = rng.sample(by_bucket[bucket], 2)
        left_split = assignments[left["component_id"]]
        right_split = assignments[right["component_id"]]
        if left_split == right_split:
            continue
        before = state_cost(
            left_split,
            states[left_split]["rows"],
            states[left_split]["strata"],
            total_rows,
            totals,
        ) + state_cost(
            right_split,
            states[right_split]["rows"],
            states[right_split]["strata"],
            total_rows,
            totals,
        )
        left_rows = states[left_split]["rows"] - left["row_count"] + right["row_count"]
        right_rows = states[right_split]["rows"] - right["row_count"] + left["row_count"]
        left_strata = swapped(states[left_split]["strata"], left["strata"], right["strata"])
        right_strata = swapped(states[right_split]["strata"], right["strata"], left["strata"])
        after = state_cost(
            left_split, left_rows, left_strata, total_rows, totals
        ) + state_cost(right_split, right_rows, right_strata, total_rows, totals)
        if after + 1e-12 >= before:
            continue
        assignments[left["component_id"]] = right_split
        assignments[right["component_id"]] = left_split
        states[left_split] = {"rows": left_rows, "strata": left_strata}
        states[right_split] = {"rows": right_rows, "strata": right_strata}
    return assignments


def split_diagnostics(
    components: list[dict[str, Any]], assignments: dict[str, str]
) -> tuple[dict[str, Any], dict[str, float], dict[str, float]]:
    total_rows = sum(component["row_count"] for component in components)
    totals = total_strata(components)
    summary: dict[str, Any] = {}
    label_errors: dict[str, float] = {}
    slice_errors: dict[str, float] = {}
    for split in SPLIT_ORDER:
        selected = [
            component
            for component in components
            if assignments[component["component_id"]] == split
        ]
        rows = [record for component in selected for record in component["records"]]
        strata = total_strata(selected)
        label_counts = {label: strata[label] for label in LABEL_ORDER}
        label_component_counts = {
            label: sum(component["strata"][label] > 0 for component in selected)
            for label in LABEL_ORDER
        }
        summary[split] = {
            "rows": len(rows),
            "row_fraction": round(len(rows) / total_rows, 6),
            "components": len(selected),
            "label_counts": label_counts,
            "label_component_counts": label_component_counts,
            "neutral": strata["neutral"],
            "cardinality_1": strata["cardinality_1"],
            "cardinality_2": strata["cardinality_2"],
            "conflicting_components": sum(
                component["conflicting_labels"] for component in selected
            ),
            "duplicate_rows": strata["duplicate_rows"],
            "duplicate_components": strata["duplicate_components"],
            "conflicting_duplicate_rows": strata["conflicting_duplicate_rows"],
        }
        fraction = SPLIT_FRACTIONS[split]
        for label in LABEL_ORDER:
            label_errors[f"{split}:{label}"] = abs(
                strata[label] / totals[label] - fraction
            )
        for name in BALANCE_SLICES:
            slice_errors[f"{split}:{name}"] = abs(
                strata[name] / totals[name] - fraction
            )
    return summary, label_errors, slice_errors


def assert_source_expectations(records: list[dict[str, Any]], duplicate_audit: dict[str, Any]) -> None:
    counts = {
        label: sum(record["labels"][index] for record in records)
        for index, label in enumerate(LABEL_ORDER)
    }
    cardinality = Counter(record["label_cardinality"] for record in records)
    if counts != EXPECTED_LABEL_COUNTS:
        raise RuntimeError(f"Frozen label counts changed: {counts}")
    if dict(sorted(cardinality.items())) != EXPECTED_CARDINALITY_COUNTS:
        raise RuntimeError(f"Frozen cardinality counts changed: {dict(cardinality)}")
    expected_duplicate = {
        "exact": {
            "unique_keys": 4_687,
            "duplicate_components": 99,
            "duplicate_rows": 212,
            "extra_duplicate_rows": 113,
            "maximum_component_size": 7,
            "conflicting_components": 26,
            "conflicting_rows": 52,
        },
        "normalized": {
            "unique_keys": 4_681,
            "duplicate_components": 99,
            "duplicate_rows": 218,
            "extra_duplicate_rows": 119,
            "maximum_component_size": 10,
            "conflicting_components": 26,
            "conflicting_rows": 52,
        },
        "connected_components": 4_681,
        "conflicting_connected_components": 26,
        "conflicting_connected_component_rows": 52,
    }
    if duplicate_audit != expected_duplicate:
        raise RuntimeError(f"Frozen duplicate audit changed: {duplicate_audit}")


def assert_split_acceptance(
    components: list[dict[str, Any]],
    assignments: dict[str, str],
    summary: dict[str, Any],
    label_errors: dict[str, float],
    slice_errors: dict[str, float],
) -> None:
    if set(assignments) != {component["component_id"] for component in components}:
        raise RuntimeError("Split assignment does not cover every component exactly once")
    if set(assignments.values()) != set(SPLIT_ORDER):
        raise RuntimeError("Split assignment has missing or unknown split names")
    row_error = max(
        abs(summary[split]["row_fraction"] - SPLIT_FRACTIONS[split])
        for split in SPLIT_ORDER
    )
    if row_error > 0.005:
        raise RuntimeError(f"Split row ratio tolerance failed: {row_error:.6f}")
    for split in SPLIT_ORDER:
        missing = [
            label for label, count in summary[split]["label_counts"].items() if count <= 0
        ]
        if missing:
            raise RuntimeError(f"Missing positive labels in {split}: {missing}")
    if max(label_errors.values(), default=0.0) > 0.05:
        raise RuntimeError(
            f"Label allocation tolerance failed: {max(label_errors.values()):.6f}"
        )
    if max(slice_errors.values(), default=0.0) > 0.03:
        raise RuntimeError(
            f"Balance-slice allocation tolerance failed: {max(slice_errors.values()):.6f}"
        )


def write_jsonl(path: Path, values: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    with temporary.open("w", encoding="utf-8", newline="\n") as handle:
        for value in values:
            handle.write(stable_json(value))
            handle.write("\n")
    temporary.replace(path)


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def file_record(path: Path) -> dict[str, Any]:
    return {
        "path": relative(path),
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def numeric_summary(values: Iterable[int]) -> dict[str, int | float]:
    ordered = sorted(values)
    return {
        "min": ordered[0],
        "median": round(statistics.median(ordered), 6),
        "mean": round(statistics.fmean(ordered), 6),
        "p95": ordered[min(len(ordered) - 1, int(0.95 * (len(ordered) - 1)))],
        "max": ordered[-1],
    }


def git_state() -> dict[str, Any]:
    commit = subprocess.run(
        ["git", "rev-parse", "HEAD"],
        cwd=PROJECT_ROOT,
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()
    dirty = subprocess.run(
        ["git", "status", "--porcelain"],
        cwd=PROJECT_ROOT,
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()
    return {"commit": commit, "worktree_dirty": bool(dirty)}


def build(
    source: Path,
    private_root: Path,
    public_report: Path,
    public_manifest: Path,
    public_split_index: Path,
) -> dict[str, Any]:
    source_sha256 = sha256_file(source)
    if source_sha256 != EXPECTED_SOURCE_SHA256:
        raise RuntimeError(f"Source SHA-256 mismatch: {source_sha256}")
    if private_root.exists():
        raise FileExistsError(f"Refusing to overwrite immutable output: {private_root}")

    records = load_records(source)
    components, duplicate_audit = build_components(records)
    assert_source_expectations(records, duplicate_audit)
    assignments = allocate_splits(components)
    summary, label_errors, slice_errors = split_diagnostics(components, assignments)
    assert_split_acceptance(components, assignments, summary, label_errors, slice_errors)

    split_records: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for component in components:
        split = assignments[component["component_id"]]
        for record in component["records"]:
            record["split"] = split
            split_records[split].append(record)
    for split in SPLIT_ORDER:
        split_records[split].sort(key=lambda value: value["sample_id"])

    private_root.parent.mkdir(parents=True, exist_ok=True)
    private_root.parent.chmod(0o700)
    temporary_root = private_root.with_name(f".{private_root.name}.tmp-{os.getpid()}")
    if temporary_root.exists():
        shutil.rmtree(temporary_root)
    temporary_root.mkdir(parents=True)
    temporary_root.chmod(0o700)

    def labeled_record(record: dict[str, Any]) -> dict[str, Any]:
        return {
            "schema_version": SCHEMA_VERSION,
            "protocol_id": PROTOCOL_ID,
            "sample_id": record["sample_id"],
            "component_id": record["component_id"],
            "text": record["text"],
            "labels": list(record["labels"]),
            "neutral": record["neutral"],
            "label_cardinality": record["label_cardinality"],
        }

    train_output = [labeled_record(record) for record in split_records["train"]]
    validation_output = [
        labeled_record(record) for record in split_records["validation"]
    ]
    test_inputs = [
        {
            "schema_version": SCHEMA_VERSION,
            "protocol_id": PROTOCOL_ID,
            "sample_id": record["sample_id"],
            "component_id": record["component_id"],
            "text": record["text"],
        }
        for record in split_records["test"]
    ]
    test_labels = [
        {
            "schema_version": "so-emotion-c0-test-label-v1",
            "protocol_id": PROTOCOL_ID,
            "sample_id": record["sample_id"],
            "labels": list(record["labels"]),
        }
        for record in split_records["test"]
    ]
    conflict_output = [
        {
            "schema_version": "so-emotion-duplicate-conflict-v1",
            "protocol_id": PROTOCOL_ID,
            "component_id": component["component_id"],
            "members": [
                {
                    "sample_id": record["sample_id"],
                    "labels": list(record["labels"]),
                }
                for record in sorted(
                    component["records"], key=lambda value: value["sample_id"]
                )
            ],
        }
        for component in components
        if component["conflicting_labels"]
    ]

    private_paths = {
        "train": temporary_root / "train.jsonl",
        "validation": temporary_root / "validation.jsonl",
        "test_inputs": temporary_root / "test.inputs.jsonl",
        "test_labels": temporary_root / "test.labels.sealed.jsonl",
        "duplicate_conflicts": temporary_root / "duplicate-conflicts.jsonl",
    }
    write_jsonl(private_paths["train"], train_output)
    write_jsonl(private_paths["validation"], validation_output)
    write_jsonl(private_paths["test_inputs"], test_inputs)
    write_jsonl(private_paths["test_labels"], test_labels)
    write_jsonl(private_paths["duplicate_conflicts"], conflict_output)
    for path in private_paths.values():
        path.chmod(0o600)

    private_artifacts_before_manifest = {}
    for name, path in private_paths.items():
        value = file_record(path)
        value["path"] = relative(private_root / path.name)
        private_artifacts_before_manifest[name] = value
    private_manifest = {
        "schema_version": "so-emotion-c0-private-manifest-v1",
        "protocol_id": PROTOCOL_ID,
        "frozen_date": FROZEN_DATE,
        "source_revision": SOURCE_REVISION,
        "source_sha256": source_sha256,
        "split_seed": SEED,
        "label_order": list(LABEL_ORDER),
        "split_summary": summary,
        "private_artifacts": private_artifacts_before_manifest,
        "test_gate": {"status": "sealed_not_authorized_for_model_access"},
    }
    write_json(temporary_root / "private-manifest.json", private_manifest)
    (temporary_root / "private-manifest.json").chmod(0o600)
    temporary_root.replace(private_root)

    private_artifacts = {
        name: file_record(private_root / path.name)
        for name, path in private_paths.items()
    }
    private_artifacts["private_manifest"] = file_record(
        private_root / "private-manifest.json"
    )

    split_index = sorted(
        (
            {
                "schema_version": "so-emotion-c0-split-index-v1",
                "protocol_id": PROTOCOL_ID,
                "sample_id": record["sample_id"],
                "component_id": record["component_id"],
                "split": record["split"],
            }
            for record in records
        ),
        key=lambda value: value["sample_id"],
    )
    write_jsonl(public_split_index, split_index)

    report = {
        "schema_version": "so-emotion-c0-construction-report-v1",
        "protocol_id": PROTOCOL_ID,
        "status": "constructed_awaiting_independent_verification",
        "frozen_date": FROZEN_DATE,
        "source": {
            "repository": "https://github.com/collab-uniba/EmotionDatasetMSR18",
            "revision": SOURCE_REVISION,
            "license_status": "no_standard_license_file; upstream README requests citation",
            "workbook": file_record(source),
            "sheets": [sheet for _, sheet in LABEL_SHEETS],
            "rows": len(records),
        },
        "construction": {
            "label_order": list(LABEL_ORDER),
            "gold_rule": "at_least_two_of_three_rater_marks",
            "sheet_alignment_mismatches": 0,
            "majority_gold_mismatches": 0,
            "label_counts": EXPECTED_LABEL_COUNTS,
            "cardinality_counts": {
                str(key): value for key, value in EXPECTED_CARDINALITY_COUNTS.items()
            },
            "emotion_present_rows": sum(not record["neutral"] for record in records),
            "neutral_rows": sum(record["neutral"] for record in records),
            "paper_release_conflict": (
                "The paper text reports 1,959 emotion and 2,841 neutral instances; "
                "the pinned workbook reconstructs the reverse, consistent with per-label "
                "supports and 133 two-label rows."
            ),
            "text_length_codepoints": numeric_summary(
                len(record["text"]) for record in records
            ),
        },
        "duplicates": duplicate_audit,
        "split": {
            "seed": SEED,
            "target_fractions": SPLIT_FRACTIONS,
            "unit": "exact_and_nfkc_casefold_whitespace_duplicate_component",
            "stratified_dimensions": list(STRATA_ORDER),
            "summary": summary,
            "maximum_row_fraction_error": round(
                max(
                    abs(summary[split]["row_fraction"] - SPLIT_FRACTIONS[split])
                    for split in SPLIT_ORDER
                ),
                6,
            ),
            "maximum_label_allocation_error": round(
                max(label_errors.values(), default=0.0), 6
            ),
            "maximum_balance_slice_allocation_error": round(
                max(slice_errors.values(), default=0.0), 6
            ),
        },
        "low_support": {
            "label": "surprise",
            "positive_rows": EXPECTED_LABEL_COUNTS["surprise"],
            "positive_components": sum(
                component["strata"]["surprise"] > 0 for component in components
            ),
            "controls": [
                "retain in six-label Macro-F1",
                "no label-specific threshold selection",
                "report per-class metrics, component bootstrap and three-seed variation",
                "report five-label sensitivity without using it for model selection",
            ],
        },
        "test_gate": {
            "status": "sealed_not_authorized_for_model_access",
            "inputs_contain_gold_fields": False,
            "labels_gitignored": True,
            "test_inputs_sha256": private_artifacts["test_inputs"]["sha256"],
            "test_labels_sha256": private_artifacts["test_labels"]["sha256"],
        },
        "public_privacy": {
            "split_index_fields": [
                "schema_version",
                "protocol_id",
                "sample_id",
                "component_id",
                "split",
            ],
            "row_text_published": False,
            "row_labels_published": False,
            "upstream_coordinates_published": False,
        },
        "reproducibility": {
            "builder": file_record(Path(__file__)),
            "git": git_state(),
        },
        "limitations": [
            "The release coordinates are not verified Stack Overflow post or thread IDs.",
            "The split is duplicate-component-disjoint, not thread-disjoint.",
            "Twenty-six duplicate components contain conflicting label vectors.",
            "Surprise has only 45 positive rows across 43 duplicate components.",
            "The repository has a citation request but no standard data license file.",
            "The workbook and paper text disagree on emotion-present versus neutral counts.",
        ],
        "claim_boundary": (
            "This is a data-construction result only; no model was trained, evaluated "
            "or selected."
        ),
    }
    write_json(public_report, report)

    manifest = {
        "schema_version": "so-emotion-c0-task-manifest-v1",
        "protocol_id": PROTOCOL_ID,
        "status": "constructed_awaiting_independent_verification",
        "source": {
            "revision": SOURCE_REVISION,
            "sha256": source_sha256,
            "license_status": "no_standard_license_file",
        },
        "counts": {
            "rows": len(records),
            "components": len(components),
            "split_rows": {
                split: summary[split]["rows"] for split in SPLIT_ORDER
            },
        },
        "public_split_index": file_record(public_split_index),
        "public_report": file_record(public_report),
        "private_root": relative(private_root),
        "private_artifacts": private_artifacts,
        "test_gate": report["test_gate"],
    }
    write_json(public_manifest, manifest)
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--private-root", type=Path, default=DEFAULT_PRIVATE_ROOT)
    parser.add_argument("--public-report", type=Path, default=DEFAULT_PUBLIC_REPORT)
    parser.add_argument("--public-manifest", type=Path, default=DEFAULT_PUBLIC_MANIFEST)
    parser.add_argument(
        "--public-split-index", type=Path, default=DEFAULT_PUBLIC_SPLIT_INDEX
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        report = build(
            args.source.resolve(),
            args.private_root.resolve(),
            args.public_report.resolve(),
            args.public_manifest.resolve(),
            args.public_split_index.resolve(),
        )
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1
    print(
        stable_json(
            {
                "protocol_id": PROTOCOL_ID,
                "status": report["status"],
                "rows": report["source"]["rows"],
                "split_rows": {
                    split: report["split"]["summary"][split]["rows"]
                    for split in SPLIT_ORDER
                },
                "maximum_label_allocation_error": report["split"][
                    "maximum_label_allocation_error"
                ],
            }
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
